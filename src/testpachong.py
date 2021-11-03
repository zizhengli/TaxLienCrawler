import requests 
from bs4 import BeautifulSoup
import openpyxl
import time
import random

wb=openpyxl.load_workbook('2021ParkCountyTS.xlsx')
sheet=wb['ParcelNumber']
parcelNumber=[]
e = sheet.iter_rows()
cells=list(e)

for i in cells:
    parcelNumber.append(i[0].value)


wb2=openpyxl.Workbook()
sheet2=wb2.active
sheet2.title='consolidate' 

firstbatch=parcelNumber[1:30]
#firstbatch=['R0023664','R0045201']
print(firstbatch)

for i in firstbatch:

    time.sleep(random.randint(1,60))
    record = []
    record.append(i)
    
    res = requests.get('https://beacon.schneidercorp.com/Application.aspx?AppID=1085&LayerID=26739&PageTypeID=4&PageID=10843&Q=550632906&KeyValue='+i)
    soup=BeautifulSoup(res.text, 'html.parser')
    
    section1=soup.find(id='ctlBodyPane_ctl00_mSection')
    items=section1.find_all('td')
    for item in items:
        title = item.text
        print(title)
        record.append(title)
    
    section2=soup.find(id='ctlBodyPane_ctl01_mSection')
    items=section2.find_all('a')
    for item in items:
        title = item.text
        print(title)
        record.append(title)
    
    items=section2.find_all('span')
    for item in items:
        title = item.text
        print(title)
        record.append(title) 

    buildingSection=soup.find(id='ctlBodyPane_ctl02_mSection')
    if buildingSection is not None:
        record.append('building')
    else:
        record.append('land')

    section3=soup.find(id='ctlBodyPane_ctl03_mSection')
    if section3 is not None:
        body=section3.find('tbody')
        column1=body.find('th')
        print(column1.text)
        record.append(column1.text)
        items=body.find_all('td')
        for item in items:
            title=item.text
            print(title)
            record.append(title)
    else:
        record.append('none')

    section5=soup.find(id='ctlBodyPane_ctl05_mSection')
    if section5 is not None:
        items=section5.find_all('tbody')
        for item in items:
            title=item.text
            print(title)
            record.append(title)

    sheet2.append(record)

    res.close()

wb2.save('batch1.xlsx')